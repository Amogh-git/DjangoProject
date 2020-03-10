from translation import variables
from nltk.stem import WordNetLemmatizer
import os
import openpyxl

os.chdir('C://Projects//Project_M//translation')
'''
wb=openpyxl.load_workbook('database.xlsx')  # to add dictionary 
#wb.sheetnames  # acess all worksheet namesnames
s = wb["Sheet1"]
nrow=s.max_row
'''
wb = openpyxl.load_workbook(
    'C:\\Projects\\Project_M\\translation\\Reference_Files\\dictionary.xlsx')  # to add dictionary
vd = wb["verb"]
nd = wb["noun"]
adj = wb["adjective"]
adv = wb["adverb"]
pr = wb["pronoun"]
s = wb["other"]  # to be added
vd_row = vd.max_row
nd_row = nd.max_row
adj_row = adj.max_row
adv_row = adv.max_row
pr_row = pr.max_row
s_row = s.max_row


lemmatizer = WordNetLemmatizer()


def marathi_list():
    i = 0
    wnpos = lambda e: ('a' if e[0].lower() == 'j' else e[0].lower()) if e[0].lower() in ['n', 'r', 'v'] else 'n'

    while i < len(variables.rearrange):  # to translate tokens into marathi
        lex = lemmatizer.lemmatize(variables.rearrange[i], wnpos(variables.pos[i]))
        variables.lexicon.append(lex)
        if variables.pos[i].startswith('V'):
            for j in range(2, vd_row + 1):
                if lex.lower() == vd.cell(row=j, column=1).value:
                    variables.marathi.append(vd.cell(row=j, column=2).value)
                    break
                else:
                    j = j + 1
        elif variables.pos[i].startswith('N'):
            for j in range(2, nd_row + 1):
                if lex.lower() == nd.cell(row=j, column=1).value:
                    variables.marathi.append(nd.cell(row=j, column=2).value)
                    break
                else:
                    j = j + 1
        elif variables.pos[i].startswith('J'):
            for j in range(2, adj_row + 1):
                if lex.lower() == adj.cell(row=j, column=1).value:
                    variables.marathi.append(adj.cell(row=j, column=2).value)
                    break
                else:
                    j = j + 1
        elif variables.pos[i].startswith('A'):
            for j in range(2, adv_row + 1):
                if lex.lower() == adv.cell(row=j, column=1).value:
                    variables.marathi.append(adv.cell(row=j, column=2).value)
                    break
                else:
                    j = j + 1
        elif variables.pos[i].startswith('P'):
            for j in range(2, pr_row + 1):
                if lex.lower() == pr.cell(row=j, column=1).value:
                    variables.marathi.append(pr.cell(row=j, column=2).value)
                    break
                else:
                    j = j + 1
        else:
            for j in range(2, s_row + 1):
                if lex.lower() == s.cell(row=j, column=1).value:
                    variables.marathi.append(s.cell(row=j, column=2).value)
                    break
                else:
                    j = j + 1

        i = i + 1

# print(variables.marathi)
